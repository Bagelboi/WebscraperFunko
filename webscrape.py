import asyncio
from playwright.async_api import async_playwright
from playwright_stealth import Stealth
import os
import json
import openpyxl
from sys import argv

DIRECTORY = argv[1]
JSON_DIRECTORY = argv[2]
EXCEL_FILE = argv[3]
CHUNK_SIZE = int(argv[4])

MAIS_OFERTA_SPAN = ".b0Sbpc"  # seletor do botão "Mais ofertas"
OFFER_ROW = ".sh-osd__offer-row"


import re

async def getOffer(page):
    # Clica em todos os "Mais ofertas" da página
    mais_ofertas = await page.query_selector_all(MAIS_OFERTA_SPAN)
    for btn in mais_ofertas:
        try:
            await btn.click()
            await page.wait_for_timeout(250)  # tempo para expandir
        except:
            pass

    ofertas = []
    ofertas_row = await page.query_selector_all(OFFER_ROW)

    for row in ofertas_row:
        try:
            texto = await row.inner_text()
            bloco = texto.strip().split("\n")

            # Vendedor geralmente é a primeira linha (antes de "Abre em uma nova janela")
            vendedor = bloco[0].strip()

            # Extrai todos os preços no texto
            precos = re.findall(r"R\$[\s\xa0]?[\d\.\,]+", texto)
            preco = " | ".join([pr.replace("\xa0", " ").strip() for pr in precos]) if precos else "N/A"

            # Pega o primeiro <a> dentro do bloco
            a_tag = await row.query_selector("a")
            url = await a_tag.get_attribute("href") if a_tag else "N/A"

            ofertas.append((vendedor, preco, url))
        except Exception as e:
            print("Erro ao processar oferta:", e)

    return ofertas

def setupSearches():
    searches = {}
    for filename in os.listdir(DIRECTORY):
        sku = filename.split("_")[-1].split(".")[0]
        if "pos" in filename:
            idx = re.search(r"pos_(\d+)", filename).group(1)
            offers = False
        else: 
            idx = filename.split("_")[1]
            offers = True
        if not sku in searches:
            name = filename.split("_")[0]
            searches[sku]  = {
                "name": name,
                "image": os.path.abspath(os.path.join(DIRECTORY, filename)),  # absolute path
                "search": []
                }
        searches[sku]["search"].append({
                "idx":idx,
                "offers":offers,
            })
    return searches

import math

async def process_chunk(p, chunk):
    """
    Processa um chunk de SKUs em um novo contexto
    """
    collected_chunk = {}
    browser = await p.chromium.launch(headless=False)
    context = await browser.new_context()
    page = await context.new_page()

    for sku, data in chunk:
        collected_chunk[sku] = {"name": data["name"], "results": []}

        json_path = os.path.join(JSON_DIRECTORY, f"{sku}.json")
        if not os.path.exists(json_path):
            print(f"⚠️ JSON não encontrado para {sku}")
            continue

        with open(json_path, "r", encoding="utf-8") as f:
            pl = json.load(f)

        res = pl["results"][0]["content"]["results"]["organic"]

        for search in data["search"]:
            idx = search["idx"]
            offers = search["offers"]

            try:
                if offers:
                    for anuncio in res:
                        if anuncio.get("product_id") == idx:
                            url_proc = anuncio["url"].split("?q=")
                            url_proc[0] += "/offers"
                            url_proc = "?q=".join(url_proc)

                            await page.goto(url_proc)
                            await page.wait_for_timeout(2000)
                            ofertas = await getOffer(page)

                            for vendedor, preco, url in ofertas:
                                collected_chunk[sku]["results"].append((vendedor, preco, url))
                else:
                    for anuncio in res:
                        if str(anuncio.get("pos_overall")) == str(idx):
                            vendedor = anuncio["merchant"]["name"]
                            preco = anuncio.get("price_str", "N/A")
                            if "url" in anuncio["merchant"]:
                                url = anuncio["merchant"]["url"]
                            else:
                                url = anuncio.get("url", "N/A")
                            collected_chunk[sku]["results"].append((vendedor, preco, url))
            except Exception as e:
                print(f"⚠️ Erro ao processar {sku} idx {idx}: {e}")

    await context.close()
    await browser.close()
    return collected_chunk


async def doSearches(searches, chunk_size):
    """
    Processa todos os SKUs em chunks simultâneos
    """
    skus = list(searches.items())
    num_chunks = math.ceil(len(skus) / chunk_size)
    chunks = [skus[i*chunk_size:(i+1)*chunk_size] for i in range(num_chunks)]

    async with Stealth().use_async(async_playwright()) as p:
        tasks = [process_chunk(p, chunk) for chunk in chunks]
        results_list = await asyncio.gather(*tasks, return_exceptions=False)

    # junta resultados de todos os chunks
    collected = {}
    for r in results_list:
        collected.update(r)

    return collected


from openpyxl.styles import Font
from openpyxl.drawing.image import Image as XLImage

def save_to_excel(results, filename):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["SKU", "Vendedor", "Preço"])

    for sku, data in results.items():
        name = data["name"]
        image_path = data.get("image")

        # Insert the name with formatting
        row_idx = ws.max_row + 1
        ws.append([f"{name} ({sku})", "", ""])
        cell = ws.cell(row=row_idx, column=1)
        cell.font = Font(size=18, bold=True)

        # Insert image if available
        if image_path and os.path.exists(image_path):
            try:
                img = XLImage(image_path)
                img.width, img.height = 64, 64  # shrink for readability
                ws.add_image(img, f"B{row_idx}")  # put image beside the name
            except Exception as e:
                print(f"⚠️ Could not insert image for {sku}: {e}")

        # Now append offers
        for vendedor, preco, url in data["results"]:
            row_idx = ws.max_row + 1
            ws.append(["", vendedor, preco])
            vcell = ws.cell(row=row_idx, column=2)
            if url and url != "N/A":
                vcell.hyperlink = url
                vcell.style = "Hyperlink"

    wb.save(filename)



if __name__ == "__main__":
    searches = setupSearches()
    results = asyncio.run(doSearches(searches, CHUNK_SIZE))
    save_to_excel(results, EXCEL_FILE)

